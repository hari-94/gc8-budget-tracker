"""
Extracts data from 'Pilot_2026_-_Budget_Tracker.xlsx' and produces two clean CSVs:
  - budget_allocations.csv   (from the 'Budget_26' tab -> one row per category/month)
  - expenses.csv             (from 'Budget Tracker' + 'Budget Tracker (2)' tabs -> one row per transaction)

These CSVs are shaped to match the Supabase tables `budget_allocations` and `expenses`
and can be imported directly via Supabase Studio: Table Editor -> table -> Insert -> Import data from CSV.

Usage:
    python extract_from_excel.py /path/to/Pilot_2026_-_Budget_Tracker.xlsx
"""
import sys
import re
import openpyxl
import csv
from datetime import datetime

MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

def norm_code(code):
    if code is None:
        return None
    return str(code).strip()

def norm_amount(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).replace('$', '').replace(',', '').strip()
    if s in ('', '-'):
        return 0.0
    try:
        return round(float(s), 2)
    except ValueError:
        return None

def extract_budget_allocations(wb, fiscal_year, out_path):
    ws = wb['Budget_26']
    seen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, category = row[0], row[1]
        if not code or not category:
            continue
        code = norm_code(code)
        # columns: Code, Category, Total, Jan..Dec  -> indices 3..14
        for i, month_name in enumerate(MONTHS):
            col_idx = 3 + i
            if col_idx >= len(row):
                continue
            amt = norm_amount(row[col_idx])
            if amt is None:
                continue
            # The sheet repeats the same chart-of-accounts block twice further down;
            # keep only the first occurrence per (code, month) to avoid double-counting.
            key = (code, i + 1)
            if key not in seen:
                seen[key] = amt
    rows_out = [
        {'category_code': code, 'fiscal_year': fiscal_year, 'month': month, 'budgeted_amount': amt}
        for (code, month), amt in seen.items()
    ]
    with open(out_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['category_code','fiscal_year','month','budgeted_amount'])
        writer.writeheader()
        writer.writerows(rows_out)
    print(f'Wrote {len(rows_out)} budget_allocations rows -> {out_path}')

def build_name_to_code_map(wb):
    """Map lowercased/stripped category name -> code, from the Budget_26 chart of accounts."""
    ws = wb['Budget_26']
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, category = row[0], row[1]
        if code and category:
            key = str(category).strip().lower()
            mapping[key] = norm_code(code)
    return mapping

# Codes found in the transaction sheets that aren't in the main Budget_26 chart of accounts.
# These are one-off / non-standard categories - keep the code but flag them so you can
# decide whether to fold them into an existing budget line or track them separately.
EXTRA_CATEGORIES = {
    '72600-530': ('Spa Infinity', 'Other'),
    '73690':     ('Parts & Supplies', 'Other'),
    '65200-71':  ('Sales Gallery Carpet', 'Other'),
    '65200-71-WC': ('Sales Gallery Window Cleaning', 'Other'),
    '65200-102': ('Sales Guest Services', 'Other'),
    '74940-550': ('Dishwasher Room Lobby Bar (not Hskp)', 'Other'),
}

def extract_expenses(wb, sheet_names, out_path):
    name_to_code = build_name_to_code_map(wb)
    rows_out = []
    unmatched = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            notes, category, code, date, month, invoice, vendor, amount = row[:8]
            if not category or amount is None:
                continue
            code = norm_code(code)
            amt = norm_amount(amount)
            if amt is None:
                continue
            # Fill in missing/blank codes by matching the category name against the chart of accounts
            if not code:
                code = name_to_code.get(str(category).strip().lower())
            if not code:
                unmatched.append((category, amount))
                continue
            # 65200-71 was reused for two different expense types in the source sheet -
            # split "Sales Gallery Window Cleaning" onto its own code so it doesn't get
            # mixed in with "Sales Gallery Carpet" totals.
            if code == '65200-71' and str(category).strip().lower() == 'sales gallery window cleaning':
                code = '65200-71-WC'
            if isinstance(date, datetime):
                txn_date = date.strftime('%Y-%m-%d')
            else:
                txn_date = ''
            rows_out.append({
                'category_code': code,
                'vendor': (str(vendor).strip() if vendor else ''),
                'invoice_number': (str(invoice).strip() if invoice else ''),
                'txn_date': txn_date,
                'amount': amt,
                'status': 'paid',
                'notes': (str(notes).strip() if notes else '')
            })
    with open(out_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['category_code','vendor','invoice_number','txn_date','amount','status','notes'])
        writer.writeheader()
        writer.writerows(rows_out)
    print(f'Wrote {len(rows_out)} expenses rows -> {out_path}')
    if unmatched:
        print(f'WARNING: {len(unmatched)} rows had no matching category code and were skipped:')
        for cat, amt in unmatched[:20]:
            print(f'   - "{cat}" (${amt})')

def write_extra_categories_csv(out_path):
    with open(out_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['code','name','group_name','type'])
        writer.writeheader()
        for code, (name, group) in EXTRA_CATEGORIES.items():
            writer.writerow({'code': code, 'name': name, 'group_name': group, 'type': 'expense'})
    print(f'Wrote {len(EXTRA_CATEGORIES)} extra_categories rows -> {out_path}')
    print('(these are one-off codes found in the transaction log that are not in Budget_26 -')
    print(' import this CSV into the "categories" table BEFORE importing expenses.csv)')

if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'Pilot_2026_-_Budget_Tracker.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    extract_budget_allocations(wb, fiscal_year=2026, out_path='budget_allocations.csv')
    write_extra_categories_csv('extra_categories.csv')
    extract_expenses(wb, sheet_names=['Budget Tracker', 'Budget Tracker (2)'], out_path='expenses.csv')
